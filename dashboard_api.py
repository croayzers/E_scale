from __future__ import annotations

import base64
import json
import re
import threading
import uuid
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties

HOST = "0.0.0.0"
PORT = 8787
DATE_FORMAT = "%Y-%m-%d %H:%M"
MAX_BODY_SIZE = 8 * 1024 * 1024

ROOT_DIR = Path(__file__).resolve().parent
DASHBOARD_DIR = ROOT_DIR / "dashboard_sharepoint_mock"
WORKBOOK_PATH = DASHBOARD_DIR / "Escale_Dashboard_Base.xlsx"
LOGO_DIR = DASHBOARD_DIR / "logos"
QUEUE_PATH = DASHBOARD_DIR / "pending_dashboard_sync.jsonl"

LOCK = threading.Lock()
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
SPACE_RE = re.compile(r"\s+")
SLUG_RE = re.compile(r"[^a-z0-9]+")
DATA_URL_RE = re.compile(r"^data:(image/[a-zA-Z0-9.+-]+);base64,(.+)$")

EXPORT_STATUS = "Activo"
RECORD_ACTIVE = "Activo"
MISSING_INFO = "No informado"


def now_text() -> str:
    return datetime.now().strftime(DATE_FORMAT)


def normalize_text(value, fallback=MISSING_INFO, *, allow_blank=False, max_length=255):
    text = SPACE_RE.sub(" ", str(value or "").strip())
    if not text:
        return "" if allow_blank else fallback
    return text[:max_length]


def normalize_email(value):
    text = normalize_text(value)
    return text if text != MISSING_INFO and EMAIL_RE.match(text) else MISSING_INFO


def normalize_int(value, default=0):
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return default


def normalize_timestamp(value):
    if not value:
        return now_text()
    text = str(value).strip()
    for parser in (datetime.fromisoformat,):
        try:
            return parser(text.replace("Z", "+00:00")).strftime(DATE_FORMAT)
        except ValueError:
            continue
    return now_text()


def slugify(value, fallback="empresa"):
    raw = normalize_text(value, fallback=fallback).lower()
    slug = SLUG_RE.sub("-", raw).strip("-")
    return slug or fallback


def json_response(handler, status, payload):
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(raw)))
    handler.send_cors_headers()
    handler.end_headers()
    handler.wfile.write(raw)


def workbook_headers(sheet):
    headers = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(5, col).value
        if value:
            headers[str(value).strip()] = col
    return headers


def first_empty_row(sheet, anchor_col=1):
    row = 6
    while True:
        if sheet.cell(row, anchor_col).value in (None, ""):
            return row
        row += 1


def find_row(sheet, headers, key, value):
    col = headers.get(key)
    if not col or not value:
        return None

    row = 6
    while True:
        row_value = sheet.cell(row, col).value
        if row_value not in (None, "") and str(row_value).strip() == str(value).strip():
            return row
        if all(sheet.cell(row, candidate).value in (None, "") for candidate in range(1, sheet.max_column + 1)):
            return None
        row += 1


def read_cell(sheet, headers, row, key, fallback=""):
    if not row:
        return fallback
    value = sheet.cell(row, headers[key]).value
    return fallback if value is None else value


def write_cells(sheet, headers, row, values):
    for key, value in values.items():
        col = headers.get(key)
        if not col:
            continue
        sheet.cell(row, col).value = value


def ensure_workbook():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"No existe el workbook base: {WORKBOOK_PATH}")
    LOGO_DIR.mkdir(parents=True, exist_ok=True)


def read_queue():
    if not QUEUE_PATH.exists():
        return []

    entries = []
    for raw in QUEUE_PATH.read_text(encoding="utf-8").splitlines():
        raw = raw.strip()
        if not raw:
            continue
        try:
            entries.append(json.loads(raw))
        except json.JSONDecodeError:
            continue
    return entries


def append_queue_entry(company_payload, export_payload=None):
    QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "company": company_payload,
        "export": export_payload,
        "queuedAt": now_text()
    }
    with QUEUE_PATH.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def clear_queue():
    if QUEUE_PATH.exists():
        QUEUE_PATH.unlink()


def parse_logo(data_url):
    if not data_url:
        return None
    match = DATA_URL_RE.match(str(data_url).strip())
    if not match:
        return None

    mime = match.group(1).lower()
    ext = {
        "image/png": "png",
        "image/jpeg": "jpg",
        "image/jpg": "jpg",
        "image/webp": "webp",
        "image/gif": "gif"
    }.get(mime, "img")
    binary = base64.b64decode(match.group(2))
    return mime, ext, binary


def save_logo_asset(company_id, company_name, data_url):
    parsed = parse_logo(data_url)
    if not parsed:
        return None

    mime, ext, binary = parsed
    asset_id = f"logo-{company_id}"
    file_name = f"{slugify(company_name)}_{company_id}.{ext}"
    file_path = LOGO_DIR / file_name
    file_path.write_bytes(binary)

    return {
        "asset_id": asset_id,
        "file_name": file_name,
        "file_extension": ext,
        "mime_type": mime,
        "relative_path": f"logos/{file_name}"
    }


def sync_logo_sheet(workbook, company_id, company_name, company_payload, company_row, company_headers):
    logos_sheet = workbook["Logos"]
    logo_headers = workbook_headers(logos_sheet)
    existing_asset_id = str(read_cell(workbook["Empresas"], company_headers, company_row, "logo_asset_id", "")).strip()
    has_logo = bool(company_payload.get("hasLogo"))
    saved_logo = save_logo_asset(company_id, company_name, company_payload.get("logoDataUrl")) if has_logo else None

    if saved_logo:
        logo_row = find_row(logos_sheet, logo_headers, "asset_id", saved_logo["asset_id"]) or first_empty_row(logos_sheet)
        write_cells(logos_sheet, logo_headers, logo_row, {
            "asset_id": saved_logo["asset_id"],
            "company_id": company_id,
            "asset_type": "Logo",
            "file_name": saved_logo["file_name"],
            "file_extension": saved_logo["file_extension"],
            "mime_type": saved_logo["mime_type"],
            "relative_path": saved_logo["relative_path"],
            "preview_status": RECORD_ACTIVE,
            "hash_reference": "",
            "last_updated_at": now_text(),
            "notes": "Logo sincronizado desde E-scale"
        })
        return saved_logo

    if company_payload.get("hasLogo") is False and existing_asset_id:
        logo_row = find_row(logos_sheet, logo_headers, "asset_id", existing_asset_id)
        if logo_row:
            write_cells(logos_sheet, logo_headers, logo_row, {
                "preview_status": "Archivado",
                "last_updated_at": now_text(),
                "notes": "Logo archivado desde E-scale"
            })
    return None


def upsert_company(workbook, company_payload):
    sheet = workbook["Empresas"]
    headers = workbook_headers(sheet)

    company_id = normalize_text(company_payload.get("id"), allow_blank=True, max_length=80)
    email = normalize_email(company_payload.get("email"))

    if not company_id and email != MISSING_INFO:
        row = find_row(sheet, headers, "company_email", email)
        if row:
            company_id = str(read_cell(sheet, headers, row, "company_id", "")).strip()

    if not company_id:
        company_id = f"CMP-{datetime.now():%Y%m%d%H%M%S}-{uuid.uuid4().hex[:6].upper()}"

    row = find_row(sheet, headers, "company_id", company_id)
    if not row:
        row = first_empty_row(sheet)

    company_name = normalize_text(company_payload.get("name"))
    logo_info = sync_logo_sheet(workbook, company_id, company_name, company_payload, row, headers)
    current_created = read_cell(sheet, headers, row, "created_at", "")

    existing_logo_asset_id = normalize_text(read_cell(sheet, headers, row, "logo_asset_id", ""), allow_blank=True)
    existing_logo_name = normalize_text(read_cell(sheet, headers, row, "logo_file_name", ""), allow_blank=True)
    existing_logo_path = normalize_text(read_cell(sheet, headers, row, "logo_relative_path", ""), allow_blank=True)

    if company_payload.get("hasLogo") is False:
        logo_asset_id = ""
        logo_file_name = ""
        logo_relative_path = ""
    elif logo_info:
        logo_asset_id = logo_info["asset_id"]
        logo_file_name = logo_info["file_name"]
        logo_relative_path = logo_info["relative_path"]
    else:
        logo_asset_id = existing_logo_asset_id
        logo_file_name = existing_logo_name
        logo_relative_path = existing_logo_path

    company_record = {
        "company_id": company_id,
        "company_name": company_name,
        "legal_name": company_name,
        "company_email": email,
        "logo_asset_id": logo_asset_id,
        "logo_file_name": logo_file_name,
        "logo_relative_path": logo_relative_path,
        "subscription_plan_current": normalize_text(company_payload.get("subscriptionPlan")),
        "subscription_status_current": normalize_text(company_payload.get("subscriptionStatus")),
        "default_event_venue": normalize_text(company_payload.get("venue")),
        "phone_future": MISSING_INFO,
        "cif_future": MISSING_INFO,
        "website": MISSING_INFO,
        "country": MISSING_INFO,
        "record_status": normalize_text(company_payload.get("recordStatus"), fallback=RECORD_ACTIVE),
        "created_at": current_created or now_text(),
        "updated_at": now_text(),
        "notes": "Sincronizado desde E-scale"
    }
    write_cells(sheet, headers, row, company_record)

    return {
        "id": company_id,
        "name": company_record["company_name"],
        "email": company_record["company_email"],
        "venue": company_record["default_event_venue"],
        "logoAssetId": company_record["logo_asset_id"],
        "logoFileName": company_record["logo_file_name"],
        "logoRelativePath": company_record["logo_relative_path"],
        "subscriptionPlan": company_record["subscription_plan_current"],
        "subscriptionStatus": company_record["subscription_status_current"],
        "recordStatus": company_record["record_status"],
        "createdAt": company_record["created_at"],
        "updatedAt": company_record["updated_at"]
    }


def append_export(workbook, company_info, export_payload):
    export_sheet = workbook["Exportaciones"]
    export_headers = workbook_headers(export_sheet)
    export_row = first_empty_row(export_sheet)
    export_id = f"EXP-{datetime.now():%Y%m%d%H%M%S}-{uuid.uuid4().hex[:6].upper()}"
    export_time = normalize_timestamp(export_payload.get("capturedAt"))

    write_cells(export_sheet, export_headers, export_row, {
        "export_id": export_id,
        "company_id": company_info["id"],
        "subscription_id": "",
        "event_id": "",
        "event_name": normalize_text(export_payload.get("eventName")),
        "venue_id": "",
        "venue_name": normalize_text(export_payload.get("venueName")),
        "export_mode": normalize_text(export_payload.get("exportMode")),
        "export_timestamp": export_time,
        "total_pax": normalize_int(export_payload.get("totalPax")),
        "total_inventory_items": normalize_int(export_payload.get("totalInventoryItems")),
        "total_inventory_categories": normalize_int(export_payload.get("totalInventoryCategories")),
        "pdf_filename": normalize_text(export_payload.get("pdfFilename")),
        "export_status": EXPORT_STATUS,
        "captured_from_app": "Si",
        "notes": "Exportación registrada desde E-scale"
    })

    inventory_sheet = workbook["Inventario"]
    inventory_headers = workbook_headers(inventory_sheet)
    current_row = first_empty_row(inventory_sheet)
    inventory_lines = export_payload.get("inventoryLines") or []

    for line in inventory_lines:
        write_cells(inventory_sheet, inventory_headers, current_row, {
            "inventory_line_id": f"INV-{uuid.uuid4().hex[:10].upper()}",
            "export_id": export_id,
            "company_id": company_info["id"],
            "category": normalize_text(line.get("category")),
            "item_type": normalize_text(line.get("itemType")),
            "item_label": normalize_text(line.get("itemLabel")),
            "quantity": normalize_int(line.get("quantity")),
            "pax": normalize_int(line.get("pax")),
            "captured_at": export_time,
            "notes": "Línea generada automáticamente desde exportación"
        })
        current_row += 1

    return {
        "exportId": export_id,
        "inventoryLines": len(inventory_lines),
        "capturedAt": export_time
    }


def persist_dashboard(company_payload, export_payload=None):
    ensure_workbook()
    queued_entries = read_queue()
    current_entry = {"company": company_payload, "export": export_payload}
    entries = queued_entries + [current_entry]

    with LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

        company_info = None
        export_info = None

        try:
            for entry in entries:
                company_info = upsert_company(workbook, entry.get("company") or {})
                if entry.get("export"):
                    export_info = append_export(workbook, company_info, entry.get("export") or {})

            workbook.save(WORKBOOK_PATH)
        except PermissionError as error:
            workbook.close()
            append_queue_entry(company_payload, export_payload)
            raise RuntimeError(
                f"El Excel maestro está abierto o bloqueado. He dejado la sincronización en cola en {QUEUE_PATH.name}. Cierra el archivo y vuelve a guardar o exportar."
            ) from error
        else:
            workbook.close()
            clear_queue()

    return company_info, export_info


def flush_queue():
    queued_entries = read_queue()
    if not queued_entries:
        return {"flushed": 0}

    with LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

        try:
            for entry in queued_entries:
                company_info = upsert_company(workbook, entry.get("company") or {})
                if entry.get("export"):
                    append_export(workbook, company_info, entry.get("export") or {})
            workbook.save(WORKBOOK_PATH)
        except PermissionError as error:
            workbook.close()
            raise RuntimeError(
                f"El Excel maestro sigue abierto o bloqueado. La cola pendiente continúa guardada en {QUEUE_PATH.name}."
            ) from error
        else:
            workbook.close()
            clear_queue()

    return {"flushed": len(queued_entries)}


class DashboardHandler(BaseHTTPRequestHandler):
    server_version = "EscaleDashboardAPI/1.0"

    def send_cors_headers(self):
        origin = self.headers.get("Origin", "*")
        self.send_header("Access-Control-Allow-Origin", origin or "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_cors_headers()
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path != "/api/dashboard/health":
            json_response(self, 404, {"ok": False, "error": "Ruta no encontrada"})
            return

        json_response(self, 200, {
            "ok": True,
            "workbook": str(WORKBOOK_PATH),
            "logosDir": str(LOGO_DIR),
            "queueFile": str(QUEUE_PATH),
            "pendingEntries": len(read_queue()),
            "updatedAt": now_text()
        })

    def do_POST(self):
        parsed = urlparse(self.path)
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length > MAX_BODY_SIZE:
                json_response(self, 413, {"ok": False, "error": "Payload demasiado grande"})
                return
            payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
        except json.JSONDecodeError:
            json_response(self, 400, {"ok": False, "error": "JSON no válido"})
            return

        try:
            if parsed.path == "/api/dashboard/company":
                company_info, _ = persist_dashboard(payload.get("company") or {})
                json_response(self, 200, {
                    "ok": True,
                    "company": company_info,
                    "syncedAt": now_text()
                })
                return

            if parsed.path == "/api/dashboard/export":
                company_info, export_info = persist_dashboard(
                    payload.get("company") or {},
                    payload.get("export") or {}
                )
                json_response(self, 200, {
                    "ok": True,
                    "company": company_info,
                    "export": export_info,
                    "syncedAt": now_text()
                })
                return

            if parsed.path == "/api/dashboard/flush":
                result = flush_queue()
                json_response(self, 200, {
                    "ok": True,
                    "flush": result,
                    "syncedAt": now_text()
                })
                return

            json_response(self, 404, {"ok": False, "error": "Ruta no encontrada"})
        except Exception as error:  # noqa: BLE001
            json_response(self, 500, {"ok": False, "error": str(error)})


def main():
    ensure_workbook()
    server = ThreadingHTTPServer((HOST, PORT), DashboardHandler)
    print(f"Dashboard API disponible en http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
